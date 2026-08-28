#!/usr/bin/env python3
"""Generate model pricing Excel from migration seed data."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

models = []

# ============================================================
# OpenAI 系列
# ============================================================
openai_models = [
    # (model_name, input_price, output_price, cached_input_price, billing_type, notes)
    ("pa/gpt-5.4-pro",           24.000000, 144.000000, 2.400000, "token", "GPT-5.4 Pro"),
    ("pa/gpt-5.4",                2.000000,  12.000000, 0.200000, "token", "GPT-5.4"),
    ("pa/gpt-5.3-chat-latest",    2.000000,  12.000000, 0.200000, "token", "GPT-5.3"),
    ("pa/gpt-5.3-codex",          1.000000,   8.000000, 0.100000, "token", "GPT-5.3 Codex"),
    ("pa/gpt-5.2-codex",          1.000000,   8.000000, 0.100000, "token", "GPT-5.2 Codex"),
    ("pa/gpt-5.2",                1.000000,   8.000000, 0.100000, "token", "GPT-5.2"),
    ("pa/gpt-5.2-chat-latest",    1.000000,   8.000000, 0.100000, "token", "GPT-5.2"),
    ("pa/gpt-5.2-pro",           24.000000, 144.000000, 2.400000, "token", "GPT-5.2 Pro"),
    ("pa/gpt-5.1-codex-max",      1.000000,   8.000000, 0.100000, "token", "GPT-5.1 Codex Max"),
    ("pa/gpt-5.1-codex-mini",     0.200000,   1.600000, 0.020000, "token", "GPT-5.1 Codex Mini"),
    ("pa/gpt-5.1",                1.000000,   8.000000, 0.100000, "token", "GPT-5.1"),
    ("pa/gpt-5.1-chat-latest",    1.000000,   8.000000, 0.100000, "token", "GPT-5.1"),
    ("pa/gpt-5.1-codex",          1.000000,   8.000000, 0.100000, "token", "GPT-5.1 Codex"),
    ("pa/gpt-5-pro",             24.000000, 144.000000, 2.400000, "token", "GPT-5 Pro"),
    ("pa/gpt-5-codex",            1.000000,   8.000000, 0.100000, "token", "GPT-5 Codex"),
    ("pa/gpt-5",                  1.000000,   8.000000, 0.100000, "token", "GPT-5"),
    ("pa/gpt-5-mini",             0.200000,   1.600000, 0.020000, "token", "GPT-5 Mini"),
    ("pa/gpt-5-nano",             0.040000,   0.320000, 0.004000, "token", "GPT-5 Nano"),
    ("pa/gpt-5-chat-latest",      1.000000,   8.000000, 0.100000, "token", "GPT-5"),
    ("pa/gt-4.1",                 1.600000,   6.400000, 0.400000, "token", "GPT-4.1"),
    ("pa/gt-4.1-n",               0.080000,   0.320000, 0.020000, "token", "GPT-4.1 Nano"),
    ("pa/gt-4.1-m",               0.320000,   1.280000, 0.080000, "token", "GPT-4.1 Mini"),
    ("pa/gt-4p",                  2.000000,   8.000000, 1.000000, "token", "GPT-4o"),
    ("pa/gt-4p-m",                0.120000,   0.480000, 0.060000, "token", "GPT-4o Mini"),
    ("pa/p1",                    12.000000,  48.000000, 6.000000, "token", "o1"),
    ("pa/p1-m",                   2.400000,   9.600000, 1.200000, "token", "o1 Mini"),
    ("pa/p3-m",                   0.880000,   3.520000, 0.440000, "token", "o3 Mini"),
    ("pa/p3",                     1.600000,   6.400000, 0.800000, "token", "o3"),
    ("pa/o4-mini",                1.600000,   6.400000, 0.800000, "token", "o4 Mini"),
    ("pa/text-embedding-3-large", 0.104000,   0.000000, 0.000000, "token", "Embedding"),
]

# ============================================================
# Anthropic 系列
# ============================================================
anthropic_models = [
    ("pa/claude-sonnet-4-6",              2.400000,  12.000000, 0.240000, "token", "Claude Sonnet 4.6"),
    ("pa/claude-opus-4-6",                4.000000,  20.000000, 0.400000, "token", "Claude Opus 4.6"),
    ("pa/claude-opus-4-5-20251101",       4.000000,  20.000000, 0.400000, "token", "Claude Opus 4.5"),
    ("pa/claude-sonnet-4-5-20250929",     2.400000,  12.000000, 0.240000, "token", "Claude Sonnet 4.5"),
    ("pa/claude-sonnet-4-5-20250929-1m",  2.400000,  12.000000, 0.240000, "token", "Claude Sonnet 4.5 (1M)"),
    ("pa/claude-haiku-4-5-20251001",      0.800000,   4.000000, 0.080000, "token", "Claude Haiku 4.5"),
    ("pa/claude-opus-4-1-20250805",      12.000000,  60.000000, 1.200000, "token", "Claude Opus 4.1"),
    ("pa/cd-st-4-20250514",              2.400000,  12.000000, 0.240000, "token", "Claude Sonnet 4"),
    ("pa/cd-op-4-20250514",             12.000000,  60.000000, 1.200000, "token", "Claude Opus 4"),
    ("pa/cd-3-7-st-20250219",            2.400000,  12.000000, 0.240000, "token", "Claude 3.7 Sonnet"),
    ("pa/cd-3-5-st-20241022",            2.400000,  12.000000, 0.240000, "token", "Claude 3.5 Sonnet"),
    ("pa/cd-3-5-hk-20241022",            0.640000,   3.200000, 0.064000, "token", "Claude 3.5 Haiku"),
    ("pa/cd-3-hk-20240307",              0.200000,   1.000000, 0.020000, "token", "Claude 3 Haiku"),
]

# ============================================================
# Google Gemini 系列
# ============================================================
gemini_models = [
    ("pa/gmn-2.5-fls",                       0.240000,   2.000000, 0.060000, "token", "Gemini 2.5 Flash"),
    ("pa/gmn-2.5-pr",                        1.000000,   8.000000, 0.250000, "token", "Gemini 2.5 Pro"),
    ("pa/gmn-2.5-fls-lt",                    0.080000,   0.320000, 0.020000, "token", "Gemini 2.5 Flash Lite"),
    ("pa/gmn-2.0-fls-20250609",              0.080000,   0.320000, 0.020000, "token", "Gemini 2.0 Flash"),
    ("pa/gmn-2.0-fls-lt",                    0.080000,   0.320000, 0.020000, "token", "Gemini 2.0 Flash Lite"),
    ("pa/gmn-2.5-fls-pw-05-20",             0.240000,   2.000000, 0.060000, "token", "Gemini 2.5 Flash Preview"),
    ("pa/gmn-2.5-pr-pw-06-05",              1.000000,   8.000000, 0.250000, "token", "Gemini 2.5 Pro Preview"),
    ("pa/gmn-2.5-fls-lt-pw-06-17",          0.080000,   0.320000, 0.020000, "token", "Gemini 2.5 Flash Lite Preview"),
    ("pa/gemini-2.5-flash-lite-preview",     0.080000,   0.320000, 0.020000, "token", "Gemini 2.5 Flash Lite Preview"),
    ("pa/gemini-3-pro-preview",              1.600000,   9.600000, 0.400000, "token", "Gemini 3 Pro Preview"),
    ("pa/gemini-3-flash-preview",            0.400000,   0.800000, 0.100000, "token", "Gemini 3 Flash Preview"),
    ("pa/gemini-3.1-pro-preview",            1.600000,   9.600000, 0.400000, "token", "Gemini 3.1 Pro Preview"),
    ("pa/gemini-3.1-flash-lite-preview",     0.080000,   0.320000, 0.020000, "token", "Gemini 3.1 Flash Lite Preview"),
    ("gemini-3-pro-image-preview",           0.100000,   0.100000, 0.000000, "image", "Gemini 3 Pro 图片生成"),
    ("gemini-3.1-flash-image-preview",       0.100000,   0.100000, 0.000000, "image", "Gemini 3.1 Flash 图片生成"),
]

# ============================================================
# Grok / xAI 系列
# ============================================================
grok_models = [
    ("pa/grok-4-1-fast-non-reasoning",  0.160000,   0.800000, 0.040000, "token", "Grok 4.1 Fast"),
    ("pa/grok-4-1-fast-reasoning",      0.160000,   0.800000, 0.040000, "token", "Grok 4.1 Fast (Reasoning)"),
    ("pa/grk-4",                        2.400000,  12.000000, 0.600000, "token", "Grok 4"),
    ("pa/grok-4-fast-reasoning",        0.160000,   0.800000, 0.040000, "token", "Grok 4 Fast (Reasoning)"),
    ("pa/grok-4-fast-non-reasoning",    0.160000,   0.800000, 0.040000, "token", "Grok 4 Fast"),
    ("pa/grok-code-fast-1",             0.160000,   0.400000, 0.040000, "token", "Grok Code Fast 1"),
    ("pa/grk-3",                        2.400000,  12.000000, 0.600000, "token", "Grok 3"),
    ("pa/grok-3-mini",                  0.240000,   0.400000, 0.060000, "token", "Grok 3 Mini"),
]

# ============================================================
# 豆包 Doubao 系列
# ============================================================
doubao_models = [
    ("pa/doubao-seed-1-8-251228",                0.089000,   0.222000, 0.009000, "token", "Doubao Seed 1.8"),
    ("pa/doubao-seed-1.6",                       0.089000,   0.222000, 0.009000, "token", "Doubao Seed 1.6"),
    ("pa/doubao-seed-1.6-thinking",              0.089000,   0.222000, 0.009000, "token", "Doubao Seed 1.6 Thinking"),
    ("pa/doubao-seed-1.6-flash",                 0.044000,   0.111000, 0.004000, "token", "Doubao Seed 1.6 Flash"),
    ("pa/doubao-1-5-pro-32k-250115",             0.089000,   0.222000, 0.009000, "token", "Doubao 1.5 Pro 32K"),
    ("pa/doubao-1.5-pro-32k-character-250715",   0.089000,   0.222000, 0.009000, "token", "Doubao 1.5 Pro 32K Character"),
]

# ============================================================
# GLM 系列
# ============================================================
glm_models = [
    ("pa/glm-5.1",  6.000000, 24.000000, 1.300000, "token", "GLM-5.1 (分梯次: Tier1 <32K: ¥6/24/1.3, Tier2 32K-200K: ¥8/28/2)"),
]

all_groups = [
    ("OpenAI", openai_models),
    ("Anthropic", anthropic_models),
    ("Google Gemini", gemini_models),
    ("Grok / xAI", grok_models),
    ("豆包 Doubao", doubao_models),
    ("GLM 智谱", glm_models),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "模型价格总表"

# Styles
header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
group_font = Font(name="微软雅黑", bold=True, size=11, color="1E3A5F")
group_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
data_font = Font(name="Consolas", size=10)
data_font_name = Font(name="微软雅黑", size=10)
price_font = Font(name="Consolas", size=10)
border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
image_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

headers = [
    "模型名称 (API ID)",
    "简称/说明",
    "计费类型",
    "输入价格\n(¥/百万tokens)",
    "输出价格\n(¥/百万tokens)",
    "缓存读取价格\n(¥/百万tokens)",
    "状态",
]

col_widths = [42, 30, 10, 18, 18, 18, 10]

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for col_idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 40

row = 2
for group_name, group_models in all_groups:
    # Group header row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f"  {group_name}")
    cell.font = group_font
    cell.fill = group_fill
    cell.alignment = left_align
    cell.border = border
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = group_fill
        ws.cell(row=row, column=c).border = border
    ws.row_dimensions[row].height = 28
    row += 1

    for model_name, input_p, output_p, cached_p, billing_type, notes in group_models:
        is_image = billing_type == "image"
        fill = image_fill if is_image else PatternFill()

        ws.cell(row=row, column=1, value=model_name).font = data_font
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value=notes).font = data_font_name
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = border

        bt_display = "按张计费" if is_image else "Token计费"
        ws.cell(row=row, column=3, value=bt_display).font = data_font_name
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = border

        if is_image:
            ws.cell(row=row, column=4, value=f"¥{input_p:.4f}/张 (1K/2K)").font = price_font
            ws.cell(row=row, column=5, value=f"¥{output_p:.4f}/张 (4K)").font = price_font
            ws.cell(row=row, column=6, value="—").font = price_font
        else:
            ws.cell(row=row, column=4, value=input_p).font = price_font
            ws.cell(row=row, column=4).number_format = '¥#,##0.000000'
            ws.cell(row=row, column=5, value=output_p).font = price_font
            ws.cell(row=row, column=5).number_format = '¥#,##0.000000'
            ws.cell(row=row, column=6, value=cached_p).font = price_font
            ws.cell(row=row, column=6).number_format = '¥#,##0.000000'

        for c in range(4, 7):
            ws.cell(row=row, column=c).alignment = center_align
            ws.cell(row=row, column=c).border = border

        ws.cell(row=row, column=7, value="启用").font = Font(name="微软雅黑", size=10, color="16A34A")
        ws.cell(row=row, column=7).alignment = center_align
        ws.cell(row=row, column=7).border = border

        if is_image:
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = fill

        row += 1

# Summary row
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
total = sum(len(g[1]) for g in all_groups)
cell = ws.cell(row=row, column=1, value=f"共计 {total} 个模型 | 价格单位: ¥人民币/百万tokens (图片模型为 ¥/张) | 数据来源: 数据库迁移种子文件")
cell.font = Font(name="微软雅黑", size=9, color="6B7280", italic=True)
cell.alignment = left_align

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:G{row - 1}"

output_path = "/root/llm-gateway/模型价格表.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print(f"Total models: {total}")
